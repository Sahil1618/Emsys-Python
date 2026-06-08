"""
Run this from your Python project folder:
    python generate_blockwise_excel.py

It reads the saved JSON from wbes_schedules/ and produces a block-wise Excel
matching the WBES portal layout (96 time blocks x schedule types per plant).
"""

import json
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
QCA       = "EESPL_QCA_BKN"
DATE      = "06-06-2026"
JSON_GLOB = f"wbes_schedules/{QCA}/{DATE}/schedule_rev_*.json"

# Time block labels (96 blocks of 15 min)
def time_label(block: int) -> str:
    mins = (block - 1) * 15
    h1, m1 = divmod(mins, 60)
    h2, m2 = divmod(mins + 15, 60)
    return f"{h1:02d}:{m1:02d}-{h2:02d}:{m2:02d}"

BLOCKS = list(range(1, 97))
TIME_LABELS = [time_label(b) for b in BLOCKS]

# ── Load JSON ────────────────────────────────────────────────────────────────
files = sorted(glob.glob(JSON_GLOB))
if not files:
    raise FileNotFoundError(f"No JSON found matching: {JSON_GLOB}")
json_path = files[-1]  # latest revision
print(f"Loading: {json_path}")

with open(json_path, encoding="utf-8") as f:
    raw = json.load(f)

# Handle both wrapped {"status":..,"data":..} and raw API response
if "data" in raw:
    api_data = raw["data"]
elif "ResponseBody" in raw:
    api_data = raw
else:
    raise ValueError("Unexpected JSON structure")

rb       = api_data.get("ResponseBody", {}) or {}
revision = rb.get("FullSchdRevisionNo", "?")
groups   = rb.get("GroupWiseDataList", []) or []

# ── Parse block-wise rows ────────────────────────────────────────────────────
# Each row: plant, schedule_type, seller, buyer, approval, [96 block values]
rows = []
for group in groups:
    acronym = group.get("Acronym", "")
    for sched in group.get("FullschdList", []) or []:
        fsd = sched.get("FullScheduleData", {}) or {}
        amounts = []
        for val in fsd.values():
            if isinstance(val, dict) and "SchdAmount" in val:
                amounts = val["SchdAmount"] or []
                break
        # Pad/trim to exactly 96
        amounts = (amounts + [0.0] * 96)[:96]
        rows.append({
            "plant":    acronym,
            "type":     sched.get("EnergyScheduleTypeName", ""),
            "seller":   sched.get("SellerAcronym", ""),
            "buyer":    sched.get("BuyerAcronym", ""),
            "approval": sched.get("ApprovalNo", ""),
            "blocks":   amounts,
        })

print(f"Parsed {len(rows)} schedule rows, Rev {revision}")

# ── Build Excel ──────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Block-wise (wide format, one row per schedule) ─────────────────
ws1 = wb.active
ws1.title = "Block-Wise (Wide)"

DARK   = "1F3864"
MID    = "2E75B6"
WHITE  = "FFFFFF"
ALT    = "EBF3FB"
GREEN  = "E2EFDA"
ORANGE = "FCE4D6"
YELLOW = "FFF2CC"

thin   = Side(style="thin", color="B8B8B8")
bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_cell(ws, row, col, val, bg=MID, fg=WHITE, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, size=9, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border    = bdr
    return c

def dat_cell(ws, row, col, val, bg=WHITE, num=False, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right" if num else "left",
                            vertical="center")
    c.border    = bdr
    if num:
        c.number_format = "0.00"
    return c

# Title
ws1.merge_cells("A1:CV1")
t = ws1["A1"]
t.value     = f"WBES Block-Wise Schedule  |  {QCA}  |  {DATE}  |  Rev {revision}"
t.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
t.fill      = PatternFill("solid", fgColor=DARK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 26

# Fixed headers (cols 1-6)
fixed_hdrs = ["Plant Acronym", "Schedule Type", "Seller", "Buyer", "Approval No", "Daily Total\n(MW)"]
for ci, h in enumerate(fixed_hdrs, 1):
    hdr_cell(ws1, 2, ci, h)

# Block headers (cols 7 onwards)
for bi, (blk, tlbl) in enumerate(zip(BLOCKS, TIME_LABELS)):
    col = 7 + bi
    hdr_cell(ws1, 2, col, f"B{blk}\n{tlbl}", bg="264478")

ws1.row_dimensions[2].height = 30

# Data rows
for ri, row in enumerate(rows):
    excel_row = ri + 3
    bg = ALT if ri % 2 == 0 else WHITE
    type_bg = GREEN if row["type"] == "AS" else ORANGE if "OA" in row["type"] else bg

    dat_cell(ws1, excel_row, 1, row["plant"],    bg=bg,      bold=True)
    dat_cell(ws1, excel_row, 2, row["type"],     bg=type_bg, bold=True)
    dat_cell(ws1, excel_row, 3, row["seller"],   bg=bg)
    dat_cell(ws1, excel_row, 4, row["buyer"],    bg=bg)
    dat_cell(ws1, excel_row, 5, row["approval"], bg=bg)
    # Daily total formula
    c = ws1.cell(row=excel_row, column=6,
                 value=f"=SUM(G{excel_row}:CV{excel_row})")
    c.font          = Font(name="Arial", size=9, bold=True)
    c.fill          = PatternFill("solid", fgColor=YELLOW)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = bdr
    c.number_format = "#,##0.00"

    for bi, val in enumerate(row["blocks"]):
        dat_cell(ws1, excel_row, 7 + bi, val, bg=bg, num=True)

# Total row
total_row = len(rows) + 3
ws1.merge_cells(f"A{total_row}:E{total_row}")
tc = ws1[f"A{total_row}"]
tc.value     = "TOTAL (ALL SCHEDULES)"
tc.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
tc.fill      = PatternFill("solid", fgColor=DARK)
tc.alignment = Alignment(horizontal="right", vertical="center")
tc.border    = bdr

for col in range(6, 7 + 96):
    c = ws1.cell(row=total_row, column=col,
                 value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{total_row-1})")
    c.font          = Font(name="Arial", bold=True, size=9, color=WHITE)
    c.fill          = PatternFill("solid", fgColor=DARK)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = bdr
    c.number_format = "#,##0.00"

# Column widths
ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 24
ws1.column_dimensions["F"].width = 13
for bi in range(96):
    ws1.column_dimensions[get_column_letter(7 + bi)].width = 8

ws1.freeze_panes = "G3"
ws1.row_dimensions[1].height = 26

# ── Sheet 2: Plant Summary (pivot by plant + block) ─────────────────────────
ws2 = wb.create_sheet("Plant Summary")

# Collect unique plants
plants = list(dict.fromkeys(r["plant"] for r in rows))

ws2.merge_cells(f"A1:{get_column_letter(2 + 96)}1")
t2 = ws2["A1"]
t2.value     = f"Plant-wise Block Summary  |  {QCA}  |  {DATE}  |  Rev {revision}"
t2.font      = Font(name="Arial", bold=True, size=12, color=WHITE)
t2.fill      = PatternFill("solid", fgColor=DARK)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

hdr_cell(ws2, 2, 1, "Plant")
hdr_cell(ws2, 2, 2, "Daily Total\n(MW)")
for bi, (blk, tlbl) in enumerate(zip(BLOCKS, TIME_LABELS)):
    hdr_cell(ws2, 2, 3 + bi, f"B{blk}\n{tlbl}", bg="264478")
ws2.row_dimensions[2].height = 30

for pi, plant in enumerate(plants):
    excel_row = pi + 3
    bg = ALT if pi % 2 == 0 else WHITE
    plant_rows = [r for r in rows if r["plant"] == plant]
    block_totals = [sum(r["blocks"][bi] for r in plant_rows) for bi in range(96)]

    dat_cell(ws2, excel_row, 1, plant, bg=bg, bold=True)
    c = ws2.cell(row=excel_row, column=2,
                 value=f"=SUM(C{excel_row}:{get_column_letter(2+96)}{excel_row})")
    c.font = Font(name="Arial", size=9, bold=True)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = bdr
    c.number_format = "#,##0.00"
    for bi, val in enumerate(block_totals):
        dat_cell(ws2, excel_row, 3 + bi, round(val, 2), bg=bg, num=True)

# Total row
tr2 = len(plants) + 3
ws2.merge_cells(f"A{tr2}:A{tr2}")
tc2 = ws2.cell(row=tr2, column=1, value="TOTAL")
tc2.font = Font(name="Arial", bold=True, size=9, color=WHITE)
tc2.fill = PatternFill("solid", fgColor=DARK)
tc2.alignment = Alignment(horizontal="right")
tc2.border = bdr
for col in range(2, 3 + 96):
    c = ws2.cell(row=tr2, column=col,
                 value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tr2-1})")
    c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="right")
    c.border = bdr
    c.number_format = "#,##0.00"

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 13
for bi in range(96):
    ws2.column_dimensions[get_column_letter(3 + bi)].width = 8
ws2.freeze_panes = "C3"

# ── Save ─────────────────────────────────────────────────────────────────────
out = f"WBES_BlockWise_{QCA}_{DATE.replace('-','')}_Rev{revision}.xlsx"
wb.save(out)
print(f"\n✅  Saved: {out}")
