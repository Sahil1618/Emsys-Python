# """Export WBES schedule rows to a plant-wise Excel file (one sheet per plant)."""

# import os
# from collections import defaultdict
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.utils import get_column_letter

# # ── colour palette ─────────────────────────────────────────────────────────
# _HDR_DARK  = "1F4E79"   # dark blue  – title bar
# _HDR_MID   = "2E75B6"   # mid blue   – meta field labels
# _BLOCK_HDR = "BDD7EE"   # light blue – block label column
# _AS_BG     = "E2EFDA"   # light green – AS rows
# _OA_BG     = "FCE4D6"   # light orange – OA_REMC rows
# _NET_BG    = "D9D2E9"   # light purple – Net Schedule column
# _NET_HDR   = "7030A0"   # purple – Net Schedule header
# _WHITE     = "FFFFFF"

# _thin  = Side(style="thin",   color="BFBFBF")
# _thick = Side(style="medium", color="595959")
# _BORDER   = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
# _BORDER_L = Border(left=_thick, right=_thin,  top=_thin,  bottom=_thin)


# def _c(ws, row, col, value="", bold=False, fg="000000", bg=None,
#        align="left", wrap=False, border=None, num_fmt=None, size=9):
#     c = ws.cell(row=row, column=col, value=value)
#     c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
#     c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
#     if bg:
#         c.fill = PatternFill("solid", start_color=bg)
#     if border:
#         c.border = border
#     if num_fmt:
#         c.number_format = num_fmt
#     return c


# def _row_bg(stype: str) -> str:
#     if stype == "AS":
#         return _AS_BG
#     if stype.startswith("OA"):
#         return _OA_BG
#     return _WHITE


# def _build_plant_sheet(ws, plant: str, plant_rows: list[dict],
#                        qca_name: str, date_str: str, revision: str) -> None:
#     """
#     Write one plant's schedules onto a worksheet.

#     External OA_REMC rows (seller != plant) are completely excluded —
#     not shown as columns and not counted in Net Schedule.
#     Net Schedule = Sum(own OA_REMC) − AS per block.
#     """
#     # Filter out external OA rows (seller != plant) entirely
#     own_rows = [
#         r for r in plant_rows
#         if not (r["type"].startswith("OA") and r.get("seller") != plant)
#     ]

#     oa_rows  = [r for r in own_rows if r["type"].startswith("OA")]
#     as_rows  = [r for r in own_rows if r["type"] == "AS"]
#     ordered  = oa_rows + as_rows
#     n_scheds = len(ordered)
#     net_col  = 2 + n_scheds

#     # ── Row 1: title ────────────────────────────────────────────────────
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=net_col)
#     c = ws.cell(row=1, column=1,
#                 value=f"{plant}  |  {qca_name}  |  {date_str}  |  Rev {revision}")
#     c.font      = Font(name="Arial", bold=True, color=_WHITE, size=11)
#     c.fill      = PatternFill("solid", start_color=_HDR_DARK)
#     c.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[1].height = 22

#     # ── Rows 2-5: meta header rows ───────────────────────────────────────
#     META = [
#         ("Schedule Type", "type"),
#         ("Sub Type",      "sub_type"),
#         ("Seller",        "seller"),
#         ("Buyer",         "buyer"),
#         ("Approval No",   "approval"),
#     ]
#     for r_off, (label, key) in enumerate(META):
#         row = 2 + r_off
#         ws.row_dimensions[row].height = 16
#         _c(ws, row, 1, label, bold=True, fg=_WHITE, bg=_HDR_MID,
#            align="center", border=_BORDER)
#         for i, srow in enumerate(ordered):
#             col = i + 2
#             bg  = _row_bg(srow["type"])
#             _c(ws, row, col, srow.get(key, ""), bg=bg,
#                align="center", wrap=True, border=_BORDER)
#         if r_off == 0:
#             _c(ws, row, net_col, "Net Schedule\n(OA − AS)",
#                bold=True, fg=_WHITE, bg=_NET_HDR,
#                align="center", wrap=True, border=_BORDER_L)
#         else:
#             _c(ws, row, net_col, "", bg=_NET_BG,
#                align="center", border=_BORDER_L)

#     # ── Row 7: Daily Total ────────────────────────────────────────────────
#     ws.row_dimensions[7].height = 22
#     _c(ws, 7, 1, "Daily Total\n(MW)", bold=True, fg=_WHITE,
#        bg=_HDR_DARK, align="center", wrap=True, border=_BORDER)

#     oa_total  = sum(r.get("daily_total_mw", 0.0) for r in oa_rows)
#     as_total  = sum(r.get("daily_total_mw", 0.0) for r in as_rows)
#     net_total = oa_total - as_total

#     for i, srow in enumerate(ordered):
#         col = i + 2
#         bg  = _row_bg(srow["type"])
#         _c(ws, 7, col, srow.get("daily_total_mw", 0.0),
#            bold=True, bg=bg, align="right",
#            border=_BORDER, num_fmt="#,##0.00")

#     _c(ws, 7, net_col, net_total,
#        bold=True, fg=_WHITE, bg=_NET_HDR,
#        align="right", border=_BORDER_L, num_fmt="#,##0.00")

#     # ── Rows 8-103: B1..B96 ──────────────────────────────────────────────
#     for b in range(96):
#         row = 8 + b
#         ws.row_dimensions[row].height = 13

#         hh,  mm  = divmod(b * 15, 60)
#         ehh, emm = divmod((b + 1) * 15, 60)
#         label = f"B{b+1}\n{hh:02d}:{mm:02d}-{ehh:02d}:{emm:02d}"
#         _c(ws, row, 1, label, bold=True, bg=_BLOCK_HDR,
#            align="center", wrap=True, border=_BORDER)

#         oa_sum = 0.0
#         as_val = 0.0

#         for i, srow in enumerate(ordered):
#             col    = i + 2
#             blocks = srow.get("blocks", [])
#             val    = blocks[b] if b < len(blocks) else 0.0
#             bg     = _row_bg(srow["type"])
#             _c(ws, row, col, val if val != 0.0 else 0,
#                bg=bg, align="right",
#                border=_BORDER, num_fmt="#,##0.00")
#             if srow["type"].startswith("OA"):
#                 oa_sum += val
#             elif srow["type"] == "AS":
#                 as_val += val

#         net_val = oa_sum - as_val
#         _c(ws, row, net_col, round(net_val, 4) if net_val != 0.0 else 0,
#            bold=True, bg=_NET_BG,
#            align="right", border=_BORDER_L, num_fmt="#,##0.00")

#     # ── Column widths & freeze ────────────────────────────────────────────
#     ws.column_dimensions["A"].width = 16
#     for col in range(2, n_scheds + 2):
#         ws.column_dimensions[get_column_letter(col)].width = 14
#     ws.column_dimensions[get_column_letter(net_col)].width = 16
#     ws.freeze_panes = ws.cell(row=8, column=2)


# def export_schedule_to_excel(
#     rows: list[dict],
#     qca_name: str,
#     date_str: str,
#     output_path: str | None = None,
# ) -> str:
#     if not rows:
#         raise ValueError("No schedule rows to export.")

#     revision = rows[0].get("revision", "NA")

#     if output_path is None:
#         safe_date   = date_str.replace("-", "")
#         output_path = os.path.join(
#             os.getcwd(),
#             f"WBES_BlockWise_{qca_name}_{safe_date}_Rev{revision}.xlsx",
#         )

#     plant_order: list[str] = []
#     by_plant: dict[str, list[dict]] = defaultdict(list)
#     for row in rows:
#         plant = row["plant"]
#         if plant not in by_plant:
#             plant_order.append(plant)
#         by_plant[plant].append(row)

#     wb = Workbook()
#     wb.remove(wb.active)

#     for plant in plant_order:
#         sheet_name = (plant[:31]
#                       .replace("/", "-").replace("\\", "-")
#                       .replace("*", "").replace("?", "")
#                       .replace("[", "").replace("]", "").replace(":", ""))
#         ws = wb.create_sheet(title=sheet_name)
#         _build_plant_sheet(ws, plant, by_plant[plant],
#                            qca_name, date_str, revision)

#     ws_sum = wb.create_sheet(title="Summary")
#     _build_summary_sheet(ws_sum, plant_order, by_plant,
#                          qca_name, date_str, revision)

#     wb.save(output_path)
#     return output_path


# def _build_summary_sheet(ws, plant_order, by_plant,
#                           qca_name, date_str, revision) -> None:
#     ws.merge_cells("A1:H1")
#     c = ws.cell(row=1, column=1,
#                 value=f"Summary  |  {qca_name}  |  {date_str}  |  Rev {revision}")
#     c.font      = Font(name="Arial", bold=True, color=_WHITE, size=11)
#     c.fill      = PatternFill("solid", start_color=_HDR_DARK)
#     c.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[1].height = 22

#     headers = ["Plant", "Schedule Type", "Sub Type", "Seller", "Buyer",
#                "Approval No", "Daily Total (MW)", "Net Schedule (MW)"]
#     for col, h in enumerate(headers, 1):
#         _c(ws, 2, col, h, bold=True, fg=_WHITE, bg=_HDR_MID,
#            align="center", border=_BORDER)
#     ws.row_dimensions[2].height = 18

#     data_row = 3
#     for plant in plant_order:
#         plant_rows = [
#             r for r in by_plant[plant]
#             if not (r["type"].startswith("OA") and r.get("seller") != plant)
#         ]
#         oa_total  = sum(r["daily_total_mw"] for r in plant_rows if r["type"].startswith("OA"))
#         as_total  = sum(r["daily_total_mw"] for r in plant_rows if r["type"] == "AS")
#         net_total = oa_total - as_total

#         for srow in plant_rows:
#             bg   = _row_bg(srow["type"])
#             vals = [
#                 plant,
#                 srow.get("type", ""),
#                 srow.get("sub_type", ""),
#                 srow.get("seller", ""),
#                 srow.get("buyer", ""),
#                 srow.get("approval", ""),
#                 srow.get("daily_total_mw", 0.0),
#                 "",
#             ]
#             for col, val in enumerate(vals, 1):
#                 nf = "#,##0.00" if col in (7, 8) else None
#                 _c(ws, data_row, col, val, bg=bg,
#                    align="right" if col >= 7 else "left",
#                    border=_BORDER, num_fmt=nf)
#             data_row += 1

#         _c(ws, data_row, 1, f"{plant} — Net", bold=True, fg=_WHITE,
#            bg=_NET_HDR, align="left", border=_BORDER)
#         for col in range(2, 8):
#             _c(ws, data_row, col, "", bg=_NET_BG, border=_BORDER)
#         _c(ws, data_row, 8, net_total, bold=True, fg=_WHITE,
#            bg=_NET_HDR, align="right", border=_BORDER, num_fmt="#,##0.00")
#         ws.row_dimensions[data_row].height = 16
#         data_row += 1

#     widths = [22, 14, 14, 18, 24, 32, 18, 18]
#     for col, w in enumerate(widths, 1):
#         ws.column_dimensions[get_column_letter(col)].width = w
#     ws.freeze_panes = "A3"

"""Export WBES schedule rows to a plant-wise Excel file (one sheet per plant)."""

import os
from collections import defaultdict
from datetime import date as _date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colour palette ─────────────────────────────────────────────────────────
_HDR_DARK  = "1F4E79"   # dark blue  – title bar
_HDR_MID   = "2E75B6"   # mid blue   – meta field labels
_BLOCK_HDR = "BDD7EE"   # light blue – block label column
_AS_BG     = "E2EFDA"   # light green – AS rows
_OA_BG     = "FCE4D6"   # light orange – OA_REMC rows
_NET_BG    = "D9D2E9"   # light purple – Net Schedule column
_NET_HDR   = "7030A0"   # purple – Net Schedule header
_WHITE     = "FFFFFF"

_thin  = Side(style="thin",   color="BFBFBF")
_thick = Side(style="medium", color="595959")
_BORDER   = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
_BORDER_L = Border(left=_thick, right=_thin,  top=_thin,  bottom=_thin)

# ── Base output root (can be overridden via env var WBES_OUTPUT_ROOT) ──────
_WBES_ROOT = os.environ.get("WBES_OUTPUT_ROOT", os.getcwd())


def _resolve_output_dir(region: str, date_str: str) -> str:
    """
    Return (and guarantee the existence of) the target directory:

        <WBES_ROOT>/
            <region>/          ← created once, never recreated
                <YYYY-MM-DD>/  ← created once per calendar day

    ``date_str`` must be in YYYY-MM-DD format (e.g. "2025-06-10").
    """
    region_clean = (
        region.strip()
              .replace("/", "-").replace("\\", "-")
              .replace(":", "").replace("*", "")
              .replace("?", "").replace('"', "")
              .replace("<", "").replace(">", "")
              .replace("|", "")
    )
    target = os.path.join(_WBES_ROOT, region_clean, date_str)
    os.makedirs(target, exist_ok=True)   # no-op if already present
    return target


def _c(ws, row, col, value="", bold=False, fg="000000", bg=None,
       align="left", wrap=False, border=None, num_fmt=None, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if border:
        c.border = border
    if num_fmt:
        c.number_format = num_fmt
    return c


def _row_bg(stype: str) -> str:
    if stype == "AS":
        return _AS_BG
    if stype.startswith("OA"):
        return _OA_BG
    return _WHITE


def _build_plant_sheet(ws, plant: str, plant_rows: list[dict],
                       qca_name: str, date_str: str, revision: str) -> None:
    """
    Write one plant's schedules onto a worksheet.

    External OA_REMC rows (seller != plant) are completely excluded —
    not shown as columns and not counted in Net Schedule.
    Net Schedule = Sum(own OA_REMC) − AS per block.
    """
    own_rows = [
        r for r in plant_rows
        if not (r["type"].startswith("OA") and r.get("seller") != plant)
    ]

    oa_rows  = [r for r in own_rows if r["type"].startswith("OA")]
    as_rows  = [r for r in own_rows if r["type"] == "AS"]
    ordered  = oa_rows + as_rows
    n_scheds = len(ordered)
    net_col  = 2 + n_scheds

    # ── Row 1: title ────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=net_col)
    c = ws.cell(row=1, column=1,
                value=f"{plant}  |  {qca_name}  |  {date_str}  |  Rev {revision}")
    c.font      = Font(name="Arial", bold=True, color=_WHITE, size=11)
    c.fill      = PatternFill("solid", start_color=_HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Rows 2-6: meta header rows ───────────────────────────────────────
    META = [
        ("Schedule Type", "type"),
        ("Sub Type",      "sub_type"),
        ("Seller",        "seller"),
        ("Buyer",         "buyer"),
        ("Approval No",   "approval"),
    ]
    for r_off, (label, key) in enumerate(META):
        row = 2 + r_off
        ws.row_dimensions[row].height = 16
        _c(ws, row, 1, label, bold=True, fg=_WHITE, bg=_HDR_MID,
           align="center", border=_BORDER)
        for i, srow in enumerate(ordered):
            col = i + 2
            bg  = _row_bg(srow["type"])
            _c(ws, row, col, srow.get(key, ""), bg=bg,
               align="center", wrap=True, border=_BORDER)
        if r_off == 0:
            _c(ws, row, net_col, "Net Schedule\n(OA − AS)",
               bold=True, fg=_WHITE, bg=_NET_HDR,
               align="center", wrap=True, border=_BORDER_L)
        else:
            _c(ws, row, net_col, "", bg=_NET_BG,
               align="center", border=_BORDER_L)

    # ── Row 7: Daily Total ────────────────────────────────────────────────
    ws.row_dimensions[7].height = 22
    _c(ws, 7, 1, "Daily Total\n(MW)", bold=True, fg=_WHITE,
       bg=_HDR_DARK, align="center", wrap=True, border=_BORDER)

    oa_total  = sum(r.get("daily_total_mw", 0.0) for r in oa_rows)
    as_total  = sum(r.get("daily_total_mw", 0.0) for r in as_rows)
    net_total = oa_total - as_total

    for i, srow in enumerate(ordered):
        col = i + 2
        bg  = _row_bg(srow["type"])
        _c(ws, 7, col, srow.get("daily_total_mw", 0.0),
           bold=True, bg=bg, align="right",
           border=_BORDER, num_fmt="#,##0.00")

    _c(ws, 7, net_col, net_total,
       bold=True, fg=_WHITE, bg=_NET_HDR,
       align="right", border=_BORDER_L, num_fmt="#,##0.00")

    # ── Rows 8-103: B1..B96 ──────────────────────────────────────────────
    for b in range(96):
        row = 8 + b
        ws.row_dimensions[row].height = 13

        hh,  mm  = divmod(b * 15, 60)
        ehh, emm = divmod((b + 1) * 15, 60)
        label = f"B{b+1}\n{hh:02d}:{mm:02d}-{ehh:02d}:{emm:02d}"
        _c(ws, row, 1, label, bold=True, bg=_BLOCK_HDR,
           align="center", wrap=True, border=_BORDER)

        oa_sum = 0.0
        as_val = 0.0

        for i, srow in enumerate(ordered):
            col    = i + 2
            blocks = srow.get("blocks", [])
            val    = blocks[b] if b < len(blocks) else 0.0
            bg     = _row_bg(srow["type"])
            _c(ws, row, col, val if val != 0.0 else 0,
               bg=bg, align="right",
               border=_BORDER, num_fmt="#,##0.00")
            if srow["type"].startswith("OA"):
                oa_sum += val
            elif srow["type"] == "AS":
                as_val += val

        net_val = oa_sum - as_val
        _c(ws, row, net_col, round(net_val, 4) if net_val != 0.0 else 0,
           bold=True, bg=_NET_BG,
           align="right", border=_BORDER_L, num_fmt="#,##0.00")

    # ── Column widths & freeze ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16
    for col in range(2, n_scheds + 2):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(net_col)].width = 16
    ws.freeze_panes = ws.cell(row=8, column=2)


def export_schedule_to_excel(
    rows: list[dict],
    qca_name: str,
    date_str: str,
    region: str,
    output_path: str | None = None,
) -> str:
    """
    Generate the plant-wise Excel workbook and save it under:

        WBES/<region>/<date_str>/WBES_BlockWise_<qca_name>_<date>_Rev<rev>.xlsx

    Parameters
    ----------
    rows        : list of schedule-row dicts (same schema as before)
    qca_name    : QCA / utility name used in the filename and title bar
    date_str    : schedule date in YYYY-MM-DD format  (e.g. "2025-06-10")
    region      : region name used as a sub-folder under WBES/
                  (e.g. "NR", "SR", "WR", "ER", "NER")
    output_path : optional full path override — if supplied the folder
                  structure is NOT created (useful for unit tests)

    Returns
    -------
    Absolute path of the saved .xlsx file.
    """
    if not rows:
        raise ValueError("No schedule rows to export.")

    revision = rows[0].get("revision", "NA")
    safe_date = date_str.replace("-", "")
    filename  = f"WBES_BlockWise_{qca_name}_{safe_date}_Rev{revision}.xlsx"

    if output_path is None:
        output_dir  = _resolve_output_dir(region, date_str)
        output_path = os.path.join(output_dir, filename)
    # If an explicit output_path was given, honour it as-is.

    plant_order: list[str] = []
    by_plant: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        plant = row["plant"]
        if plant not in by_plant:
            plant_order.append(plant)
        by_plant[plant].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    for plant in plant_order:
        sheet_name = (plant[:31]
                      .replace("/", "-").replace("\\", "-")
                      .replace("*", "").replace("?", "")
                      .replace("[", "").replace("]", "").replace(":", ""))
        ws = wb.create_sheet(title=sheet_name)
        _build_plant_sheet(ws, plant, by_plant[plant],
                           qca_name, date_str, revision)

    ws_sum = wb.create_sheet(title="Summary")
    _build_summary_sheet(ws_sum, plant_order, by_plant,
                         qca_name, date_str, revision)

    wb.save(output_path)
    return output_path


def _build_summary_sheet(ws, plant_order, by_plant,
                          qca_name, date_str, revision) -> None:
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value=f"Summary  |  {qca_name}  |  {date_str}  |  Rev {revision}")
    c.font      = Font(name="Arial", bold=True, color=_WHITE, size=11)
    c.fill      = PatternFill("solid", start_color=_HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["Plant", "Schedule Type", "Sub Type", "Seller", "Buyer",
               "Approval No", "Daily Total (MW)", "Net Schedule (MW)"]
    for col, h in enumerate(headers, 1):
        _c(ws, 2, col, h, bold=True, fg=_WHITE, bg=_HDR_MID,
           align="center", border=_BORDER)
    ws.row_dimensions[2].height = 18

    data_row = 3
    for plant in plant_order:
        plant_rows = [
            r for r in by_plant[plant]
            if not (r["type"].startswith("OA") and r.get("seller") != plant)
        ]
        oa_total  = sum(r["daily_total_mw"] for r in plant_rows if r["type"].startswith("OA"))
        as_total  = sum(r["daily_total_mw"] for r in plant_rows if r["type"] == "AS")
        net_total = oa_total - as_total

        for srow in plant_rows:
            bg   = _row_bg(srow["type"])
            vals = [
                plant,
                srow.get("type", ""),
                srow.get("sub_type", ""),
                srow.get("seller", ""),
                srow.get("buyer", ""),
                srow.get("approval", ""),
                srow.get("daily_total_mw", 0.0),
                "",
            ]
            for col, val in enumerate(vals, 1):
                nf = "#,##0.00" if col in (7, 8) else None
                _c(ws, data_row, col, val, bg=bg,
                   align="right" if col >= 7 else "left",
                   border=_BORDER, num_fmt=nf)
            data_row += 1

        _c(ws, data_row, 1, f"{plant} — Net", bold=True, fg=_WHITE,
           bg=_NET_HDR, align="left", border=_BORDER)
        for col in range(2, 8):
            _c(ws, data_row, col, "", bg=_NET_BG, border=_BORDER)
        _c(ws, data_row, 8, net_total, bold=True, fg=_WHITE,
           bg=_NET_HDR, align="right", border=_BORDER, num_fmt="#,##0.00")
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    widths = [22, 14, 14, 18, 24, 32, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"